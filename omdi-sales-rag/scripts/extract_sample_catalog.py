from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


def clean(value) -> str:
    return " ".join(str(value).replace("\n", " ").split()) if value not in (None, "") else ""


def main() -> None:
    sample_dir = Path("sample_data/yigit-aluminium")
    workbook_path = sample_dir / "product_image_catalog.xlsx"
    output_path = sample_dir / "catalog_snapshot.csv"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []
    try:
        sheet = workbook["KATALOG"]
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [clean(value) for value in row]
            if row_number <= 3 or not any(values):
                continue
            code = values[1] if len(values) > 1 else ""
            name = values[3] if len(values) > 3 else ""
            length = values[4] if len(values) > 4 else ""
            color = values[5] if len(values) > 5 else ""
            if code:
                rows.append(
                    {
                        "product_code": code,
                        "product_name": name,
                        "length": length,
                        "color": color,
                        "source_sheet": sheet.title,
                        "source_row": row_number,
                    }
                )
    finally:
        workbook.close()

    with output_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=[
                "product_code",
                "product_name",
                "length",
                "color",
                "source_sheet",
                "source_row",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {len(rows)} products to {output_path}")


if __name__ == "__main__":
    main()

