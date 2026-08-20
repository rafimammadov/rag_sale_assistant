from __future__ import annotations

import hashlib
import io
import re
from pathlib import Path

from PIL import Image, ImageOps, UnidentifiedImageError

SKU_PATTERN = re.compile(r"\b[A-Za-z]\d+(?:\.\d+)?\b")
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}


def normalize_sku(value: str) -> str:
    candidate = value.strip().upper()
    if not re.fullmatch(r"[A-Z]\d+(?:\.\d+)?", candidate):
        raise ValueError("Invalid product code.")
    return candidate


def extract_skus(value: str) -> list[str]:
    return list(dict.fromkeys(normalize_sku(match) for match in SKU_PATTERN.findall(value)))


class ProductMediaStore:
    """Extracts and resolves customer-safe product previews by exact SKU."""

    def __init__(self, data_dir: Path, *, max_images_per_sku: int = 2):
        self.root = data_dir / "product-media"
        self.max_images_per_sku = max(1, max_images_per_sku)

    def _sku_dir(self, company_id: str, sku: str) -> Path:
        return self.root / company_id / normalize_sku(sku)

    def images(self, company_id: str, sku: str) -> list[Path]:
        directory = self._sku_dir(company_id, sku)
        if not directory.exists():
            return []
        return sorted(
            path
            for path in directory.iterdir()
            if path.is_file() and path.suffix.casefold() in IMAGE_EXTENSIONS
        )

    def first_image(self, company_id: str, sku: str) -> Path | None:
        paths = self.images(company_id, sku)
        return paths[0] if paths else None

    def _save_image(
        self,
        *,
        company_id: str,
        sku: str,
        image: Image.Image,
        key: str,
    ) -> bool:
        directory = self._sku_dir(company_id, sku)
        existing = self.images(company_id, sku)
        if len(existing) >= self.max_images_per_sku:
            return False
        destination = directory / f"{key}.jpg"
        if destination.exists():
            return False

        prepared = ImageOps.exif_transpose(image)
        prepared.thumbnail((1600, 1600))
        if prepared.mode in {"RGBA", "LA"} or (
            prepared.mode == "P" and "transparency" in prepared.info
        ):
            rgba = prepared.convert("RGBA")
            background = Image.new("RGBA", rgba.size, "white")
            background.alpha_composite(rgba)
            prepared = background.convert("RGB")
        else:
            prepared = prepared.convert("RGB")

        directory.mkdir(parents=True, exist_ok=True)
        prepared.save(destination, format="JPEG", quality=88, optimize=True)
        return True

    def _extract_xlsx(self, company_id: str, path: Path, source_key: str) -> int:
        from openpyxl import load_workbook

        saved = 0
        workbook = load_workbook(path, read_only=False, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                for image_index, drawing in enumerate(worksheet._images, start=1):
                    anchor = getattr(drawing.anchor, "_from", None)
                    if anchor is None:
                        continue
                    row_number = anchor.row + 1
                    row_text = " ".join(
                        str(worksheet.cell(row_number, column).value or "")
                        for column in range(1, worksheet.max_column + 1)
                    )
                    skus = extract_skus(row_text)
                    if len(skus) != 1:
                        continue
                    try:
                        image = Image.open(io.BytesIO(drawing._data()))
                        key = f"{source_key}-r{row_number:04d}-{image_index:03d}"
                        saved += int(
                            self._save_image(
                                company_id=company_id,
                                sku=skus[0],
                                image=image,
                                key=key,
                            )
                        )
                    except (OSError, UnidentifiedImageError, ValueError):
                        continue
        finally:
            workbook.close()
        return saved

    def _extract_pdf(self, company_id: str, path: Path, source_key: str) -> int:
        from pypdf import PdfReader

        saved = 0
        reader = PdfReader(str(path))
        for page_number, page in enumerate(reader.pages, start=1):
            skus = extract_skus(page.extract_text() or "")
            if len(skus) != 1:
                continue
            for image_index, image_file in enumerate(page.images, start=1):
                if len(self.images(company_id, skus[0])) >= self.max_images_per_sku:
                    break
                try:
                    image = image_file.image
                    if image.width < 220 or image.height < 180:
                        continue
                    key = f"{source_key}-p{page_number:04d}-{image_index:03d}"
                    saved += int(
                        self._save_image(
                            company_id=company_id,
                            sku=skus[0],
                            image=image,
                            key=key,
                        )
                    )
                except (OSError, UnidentifiedImageError, ValueError):
                    continue
        return saved

    def extract_file(self, company_id: str, path: Path) -> int:
        extension = path.suffix.casefold()
        if extension not in {".pdf", ".xlsx"}:
            return 0
        source_key = hashlib.sha256(path.read_bytes()).hexdigest()[:12]
        if extension == ".xlsx":
            return self._extract_xlsx(company_id, path, source_key)
        return self._extract_pdf(company_id, path, source_key)
